"""
Helper script to create Excel test file for integration tests.
Generates a .xlsx file with sample employee data including cell comments.

singer_encodings preserves cell comments in synced records as:
    [{"text": <cell_value>, "comment": {"text": "...", "excel_author": "..."}}]
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.comments import Comment
import os

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

def create_excel_test_file():
    """Create employees.xlsx with sample employee data and cell comments."""
    excel_path = os.path.join(SCRIPT_DIR, 'employees.xlsx')

    # Create a new workbook and select the active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Employees"

    # Add headers
    headers = ['employee_id', 'first_name', 'last_name', 'department', 'hire_date', 'salary']
    ws.append(headers)

    # Add employee data rows
    employees = [
        [1, 'Alice', 'Johnson', 'Engineering', '2020-01-15', 95000],
        [2, 'Bob', 'Smith', 'Marketing', '2019-06-20', 75000],
        [3, 'Carol', 'Williams', 'Sales', '2021-03-10', 65000],
        [4, 'David', 'Brown', 'Engineering', '2018-11-05', 105000],
        [5, 'Eve', 'Davis', 'HR', '2022-02-28', 70000],
        [6, 'Frank', 'Miller', 'Engineering', '2020-07-12', 98000],
        [7, 'Grace', 'Wilson', 'Marketing', '2021-09-01', 78000],
        [8, 'Henry', 'Moore', 'Sales', '2019-04-18', 68000],
        [9, 'Ivy', 'Taylor', 'Engineering', '2021-11-22', 92000],
        [10, 'Jack', 'Anderson', 'HR', '2020-05-30', 72000],
    ]

    for employee in employees:
        ws.append(employee)

    # Add cell comments to exercise singer-encodings comment preservation.
    # singer_encodings returns commented cells as:
    #   [{"text": <value>, "comment": {"text": "...", "excel_author": "..."}}]
    ws['A2'].comment = Comment("Primary key field", "QA")
    ws['D3'].comment = Comment("Department is source-system maintained", "QA")
    ws['F2'].comment = Comment("Annual salary in USD", "QA")

    # Save the workbook
    wb.save(excel_path)

    print(f"Created {excel_path}")
    print(f"  - 10 employee records")
    print(f"  - Cell comments on A2, D3, F2")
    print(f"  - Columns: {', '.join(headers)}")


if __name__ == "__main__":
    try:
        create_excel_test_file()
        print("\nExcel test file is ready!")
    except ImportError:
        print("Error: openpyxl is not installed.")
        print("Install it with: pip install openpyxl")
    except Exception as e:
        print(f"Error creating Excel file: {e}")
